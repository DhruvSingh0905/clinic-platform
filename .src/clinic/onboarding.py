"""Onboarding pipeline: Excel/CSV import with LLM-powered column mapping.

Flow:
1. Clinician uploads spreadsheets (xlsx/csv)
2. We read headers + sample rows
3. LLM classifies data type and proposes column mappings
4. Clinician reviews/resolves conflicts
5. Data imported into platform tables
"""

import csv
import io
import json
import os
import sqlite3
import uuid
from datetime import datetime
from pathlib import Path

import anthropic

from clinic.config import ANTHROPIC_API_KEY, UPLOADS_DIR


# ── Target schema description for the LLM ──

TARGET_SCHEMA = """
You are mapping spreadsheet columns to a health/fitness clinic platform database.

Available target tables and their columns:

1. **patient** — basic patient info
   - name (TEXT), email (TEXT)

2. **training_block** — training program phases
   - name (TEXT), block_type (TEXT: exercise/rehab/maintenance/custom), start_date (DATE), end_date (DATE), notes (TEXT)

3. **nutrition_target** — daily nutrition targets
   - calories (INT), protein_g (INT), carbs_g (INT), fat_g (INT), effective_date (DATE), notes (TEXT)

4. **wearable_observation** — daily vitals/biometrics
   - metric (TEXT: weight_kg/resting_hr/hrv_rmssd/recovery_score/bp_systolic/bp_diastolic), observation_date (DATE), value_mean (REAL), unit (TEXT)

5. **metric_observation** — lab/bloodwork results
   - metric_name (TEXT: test name like "Hematocrit", "ALT", "Creatinine"), value (REAL), unit (TEXT), observation_date (DATE), flag (TEXT: high/low/normal)

6. **compound_event** — substance protocol entries (patient self-reported)
   - compound_name (TEXT), event_type (TEXT: START/STOP/DOSE_CHANGE), dose_mg (REAL), frequency (TEXT), route (TEXT: IM/oral/subQ), timestamp (DATE)

7. **recovery_note** — clinician notes
   - note_type (TEXT: assessment/plan/subjective/objective/follow_up/general), content (TEXT), created_at (DATE)
"""

MAPPING_PROMPT = """Analyze this spreadsheet data and produce a JSON mapping.

{schema}

## Spreadsheet: "{filename}"
### Headers: {headers}
### Sample rows (first 5):
{sample_rows}

## Your task:
1. Classify what type of data this is (training, nutrition, bloodwork, vitals, protocol, notes, patient_info, unknown)
2. For each column header, map it to: target_table.target_column, or "skip" if it doesn't map
3. Identify any ambiguous or problematic mappings

Respond with ONLY valid JSON in this exact format:
{{
  "data_type": "training|nutrition|bloodwork|vitals|protocol|notes|patient_info|unknown",
  "confidence": 0.0-1.0,
  "column_mappings": {{
    "OriginalColumnName": {{
      "target": "table.column" or "skip",
      "confidence": 0.0-1.0,
      "note": "optional explanation"
    }}
  }},
  "conflicts": [
    {{
      "column": "ColumnName",
      "issue": "ambiguous|unmapped|type_mismatch",
      "description": "explanation"
    }}
  ]
}}"""


def read_spreadsheet(file_path: str) -> tuple[list[str], list[list[str]], int]:
    """Read a spreadsheet and return (headers, sample_rows, total_row_count)."""
    ext = Path(file_path).suffix.lower()

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return [], [], 0
        headers = rows[0]
        data_rows = rows[1:]
        sample = [row for row in data_rows[:5]]
        return headers, sample, len(data_rows)

    elif ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel files. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        headers = [str(c) if c else f"col_{i}" for i, c in enumerate(next(rows_iter, []))]
        data_rows = []
        sample = []
        for i, row in enumerate(rows_iter):
            data_rows.append(row)
            if i < 5:
                sample.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return headers, sample, len(data_rows)

    else:
        raise ValueError(f"Unsupported file type: {ext}")


def analyze_with_llm(
    headers: list[str],
    sample_rows: list[list[str]],
    filename: str,
) -> dict:
    """Use Claude to analyze spreadsheet structure and propose column mappings."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    sample_text = "\n".join(
        [" | ".join(str(c) for c in row) for row in sample_rows]
    )

    prompt = MAPPING_PROMPT.format(
        schema=TARGET_SCHEMA,
        filename=filename,
        headers=json.dumps(headers),
        sample_rows=sample_text,
    )

    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=2000,
        messages=[{"role": "user", "content": prompt}],
    )

    text = response.content[0].text.strip()
    # Extract JSON from response (handle markdown code blocks)
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0].strip()
    elif "```" in text:
        text = text.split("```")[1].split("```")[0].strip()

    return json.loads(text)


def store_onboard_file(session_id: str, content: bytes, filename: str) -> str:
    """Store an uploaded onboarding file and return the relative path."""
    upload_dir = os.path.join(UPLOADS_DIR, "onboard", session_id)
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"{uuid.uuid4().hex[:8]}_{filename}"
    path = os.path.join(upload_dir, safe_name)
    with open(path, "wb") as f:
        f.write(content)
    return path


def execute_import(
    conn: sqlite3.Connection,
    session_id: str,
    file_id: int,
    patient_id: str,
    clinician_id: str,
) -> dict:
    """Execute the confirmed import for a single file."""
    file_row = conn.execute(
        "SELECT storage_path, mapping_json, detected_data_type FROM onboarding_file WHERE id = ?",
        (file_id,),
    ).fetchone()

    if not file_row:
        return {"error": "File not found"}

    mapping = json.loads(file_row["mapping_json"]) if file_row["mapping_json"] else {}
    col_mappings = mapping.get("column_mappings", {})
    data_type = file_row["detected_data_type"] or mapping.get("data_type", "unknown")

    headers, sample_rows, _ = read_spreadsheet(file_row["storage_path"])
    # Re-read full file for import
    ext = Path(file_row["storage_path"]).suffix.lower()
    all_rows: list[list[str]] = []

    if ext == ".csv":
        with open(file_row["storage_path"], "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            all_rows = list(reader)
    elif ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(file_row["storage_path"], read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        next(rows_iter)  # skip header
        all_rows = [[str(c) if c is not None else "" for c in row] for row in rows_iter]
        wb.close()

    imported = 0
    errors = 0

    for row in all_rows:
        try:
            row_dict = {}
            for i, header in enumerate(headers):
                if i < len(row):
                    m = col_mappings.get(header, {})
                    target = m.get("target", "skip") if isinstance(m, dict) else "skip"
                    if target != "skip":
                        table, col = target.split(".", 1)
                        if table not in row_dict:
                            row_dict[table] = {}
                        row_dict[table][col] = row[i]

            # Insert into appropriate tables
            for table, cols in row_dict.items():
                if table == "wearable_observation" and "metric" in cols:
                    conn.execute(
                        """INSERT OR IGNORE INTO wearable_observation
                           (user_id, metric, observation_date, value_mean, unit, source, reading_count)
                           VALUES (?, ?, ?, ?, ?, 'import', 1)""",
                        (patient_id, cols.get("metric", ""), cols.get("observation_date", ""), float(cols.get("value_mean", 0)), cols.get("unit", "")),
                    )
                elif table == "metric_observation" and "metric_name" in cols:
                    conn.execute(
                        """INSERT INTO metric_observation
                           (user_id, metric_loinc, value_canonical, unit_canonical, observation_date, flag, notes)
                           VALUES (?, 'IMPORT', ?, ?, ?, ?, 'Imported via onboarding')""",
                        (patient_id, float(cols.get("value", 0)), cols.get("unit", ""), cols.get("observation_date", ""), cols.get("flag", "")),
                    )
                elif table == "nutrition_target":
                    conn.execute(
                        """INSERT INTO nutrition_target (id, patient_id, clinician_id, calories, protein_g, carbs_g, fat_g, effective_date, notes)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                        (uuid.uuid4().hex, patient_id, clinician_id, int(cols.get("calories", 0)), int(cols.get("protein_g", 0)), int(cols.get("carbs_g", 0)), int(cols.get("fat_g", 0)), cols.get("effective_date", datetime.now().strftime("%Y-%m-%d")), cols.get("notes", "")),
                    )
                elif table == "training_block":
                    conn.execute(
                        """INSERT INTO training_block (id, patient_id, clinician_id, name, block_type, start_date, end_date, notes)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (uuid.uuid4().hex, patient_id, clinician_id, cols.get("name", "Imported Block"), cols.get("block_type", "exercise"), cols.get("start_date", ""), cols.get("end_date", ""), cols.get("notes", "")),
                    )
                elif table == "compound_event":
                    # Substance data imported as patient self-report
                    from clinic.compound_db import find_compound
                    comp = find_compound(cols.get("compound_name", ""))
                    compound_id = comp["id"] if comp else "unknown"
                    conn.execute(
                        """INSERT INTO compound_event (user_id, compound_id, event_type, timestamp, dose_mg, frequency, route, source_quality)
                           VALUES (?, ?, ?, ?, ?, ?, ?, 'import')""",
                        (patient_id, compound_id, cols.get("event_type", "START"), cols.get("timestamp", datetime.now().isoformat()), float(cols.get("dose_mg", 0)) if cols.get("dose_mg") else None, cols.get("frequency"), cols.get("route")),
                    )
                elif table == "recovery_note":
                    conn.execute(
                        """INSERT INTO recovery_note (id, patient_id, clinician_id, note_type, content)
                           VALUES (?, ?, ?, ?, ?)""",
                        (uuid.uuid4().hex, patient_id, clinician_id, cols.get("note_type", "note"), cols.get("content", "")),
                    )
            imported += 1
        except Exception:
            errors += 1

    conn.commit()
    return {"imported": imported, "errors": errors, "data_type": data_type}
